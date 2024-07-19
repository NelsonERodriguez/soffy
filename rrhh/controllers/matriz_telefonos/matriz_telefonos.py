from os import unlink
import django.db as db
import time

from array import array
from tempfile import TemporaryFile
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from core.functions import get_query, set_notification, insert_query, execute_query
from django.db import connection
from datetime import datetime
from django.core.files.storage import FileSystemStorage
from rrhh.models import Telefono, Telefono_descuento
import openpyxl
import json


@login_required(login_url="/login/")
def index(request):
    str_query = """
        SELECT a.*, case b.Descripcion when 'S' then 'Soltero(a)' when 'C' then 'Casado(a)' when 'U' then 'Unido(a)' end as EstadoCivil
        FROM NOVA..rrhh_datos_empleado as a
        JOIN NominaGB..EstadoCivil as b ON a.estado_civil = b.No_EstadoCivil
        WHERE bool_cambio_datos = 1
        AND bool_actualizados = 0
    """
    
    obj_datos = get_query(str_query)

    data = {
        "datos": obj_datos
    }

    return render(request, 'matriz_telefonos/matriz_telefonos.html', data)


@login_required(login_url='/login/')
def procesar(request):
    url = 'media/matriz_telefonos/'
    file_excel = request.FILES.get('filArchivo')

    ts = time.time()
    str_time_stamp = str(ts)

    if file_excel:
        fs = FileSystemStorage(location=url)
        path = fs.save("ArchivoMatriz_" + str_time_stamp + ".xlsx", file_excel)

    wb = openpyxl.load_workbook(file_excel)
    sheets = wb.sheetnames
    # Siempre vamos a buscar la página dos del documento de matriz que es subido
    worksheet = wb[sheets[1]]

    int_column_tel = None
    int_row_tel = 0
    int_column_total = None

    int_row = 0

    int_row_fin = 0

    for row in worksheet.iter_rows():
        int_row = int_row + 1
        for cell in row:
            value = str(cell.value)
            if value.find("Telefono", 0, 15) >= 0:
                int_column_tel = cell.column - 1
                int_row_tel = int_row
            if value.find("Tot.Monto", 0, 15) >= 0:
                int_column_total = cell.column - 1
            if value.find("Monto factura:", 0, 15) >= 0:
                int_row_fin = int_row

    int_row = 0
    datos = list()

    for row in worksheet.iter_rows():
        int_row = int_row + 1
        if int_row_tel < int_row < (int_row_fin - 1):
            cell_tel = row[int_column_tel]

            str_celular = str(cell_tel.value)
            if str_celular != "None":
                int_celular = int(str_celular)
                total = float(row[int_column_total].value)
                obj_telefono = Telefono.objects.filter(numero=int_celular).first()
                if obj_telefono.user_id:
                    str_user_id = str(obj_telefono.user_id)

                    str_query = """
                        SELECT TOP 1 eb.no_empleado, eb.empleado_id, em.nombre, em.apellido
                        FROM ares..empleados_base AS eb
                        JOIN ares..empleados_master AS em ON em.id = eb.empleado_id
                        JOIN NOVA..auth_user AS au ON au.empleado_id = em.id
                        WHERE au.id = %s
                        ORDER BY eb.fecha_baja, eb.empleado_id, eb.no_empleado desc
                    """ % str_user_id
                    obj_query = get_query(str_query)

                    subsidio = float(obj_telefono.subsidio)
                    pagar = total - subsidio

                    if pagar > 0:
                        pagar = pagar
                    else:
                        pagar = 0.00

                    if subsidio < total:
                        subsidio = subsidio
                    else:
                        subsidio = total

                    registro = {
                        'codigo': obj_query[0]["no_empleado"],
                        'empleado': obj_query[0]["nombre"] + " " + obj_query[0]["apellido"],
                        'telefono': str_celular,
                        'cobro': round(total, 2),
                        'subsidio': round(subsidio, 2),
                        'pagar': round(pagar, 2)
                    }

                    datos.append(registro)

                else:
                    registro = {
                        'codigo': "",
                        'empleado': "No Asignado",
                        'telefono': str_celular,
                        'cobro': round(total, 2),
                        'subsidio': 0.00,
                        'pagar': round(total, 2)
                    }

                    datos.append(registro)

    obj_json = {
        "status": True,
        "msj": "Se recorrió el archivo matriz correctamente.",
        "data": datos
    }
    unlink(url+path)

    return JsonResponse(obj_json, safe=False)


@login_required(login_url='/login/')
def guardar(request):

    obj_data = request
    bool_status = True
    str_msj = "Se guardaron los datos del archivos matriz correctamente."

    obj_request = json.loads(request.body)

    str_query = """
        SELECT TOP 1 * FROM NominaGB..Movimientos WHERE No_Periodo IN(
            SELECT No_Periodo
            FROM NominaGB..Periodos
            WHERE Fecha_Final = (DATEADD(month, ((YEAR(GETDATE()) - 1900) * 12) + MONTH(GETDATE()), -1)))
            and No_Clave = 59
    """
    obj_nomina = get_query(str_query)

    str_query = """
        SELECT TOP 1 * FROM NominaGB..Movimientos WHERE No_Periodo IN(
            SELECT No_Periodo
            FROM NominaGBF..Periodos
            WHERE Fecha_Final = (DATEADD(month, ((YEAR(GETDATE()) - 1900) * 12) + MONTH(GETDATE()), -1)))
            and No_Clave = 59
    """
    obj_nominaf = get_query(str_query)

    str_query = """
        SELECT TOP 1 * FROM NominaGB..Movimientos WHERE No_Periodo IN(
            SELECT No_Periodo
            FROM NominaGBV..Periodos
            WHERE Fecha_Final = (DATEADD(month, ((YEAR(GETDATE()) - 1900) * 12) + MONTH(GETDATE()), -1))
        )
        and No_Clave = 59
    """
    obj_nominav = get_query(str_query)

    if obj_nomina or obj_nominaf or obj_nominav:
        bool_status = False
        str_msj = "Ya se registraron los datos de este mes."
        set_notification(request, True, str_msj, "warning", "danger")
    else:
        if len(obj_request) > 0:
            for row in obj_request["descuentos"]:
                obj_telefono = Telefono.objects.filter(numero=row["telefono"]).first()
                empresa_id = obj_telefono.user.empresa_id if obj_telefono.user else None

                Telefono_descuento.objects.create(
                    telefono_id=obj_telefono.id,
                    subsidio=round(float(row["subsidio"]), 2),
                    empresa_id=empresa_id,
                    facturado=row["cobro"],
                )

            execute_query("EXEC NOVA..MovimientosTelefonos")

            set_notification(request, True, str_msj, "add_alert", "success")
        else:
            str_msj = "No hay datos para guardar."
            set_notification(request, True, str_msj, "warning", "danger")
            bool_status = False

    obj_json = {
        "status": bool_status,
        "msj": str_msj
    }

    return JsonResponse(obj_json, safe=False)
