from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from core.functions import get_query, set_notification, insert_query, execute_query
import openpyxl
import datetime as dt
import unidecode
import calendar


int_key_name_column = 1
arr_all = ['cuotas_ahorro', 'cuotas_csl', 'microcreditos', 'cuotas_figua', 'farmacia',
            'cuotas_upa', 'cuotas_otros', 'cuota_seguro', 'cuota_jornadas']
columns = [
    {
        'str_position': 'codigo_empleado',
        'apply': arr_all,
    },
    {
        'str_position': 'nombres',
        'apply': arr_all,
    },
    {
        'str_position': 'cuota',
        'apply': arr_all,
    },
]
arr_nominas = ['nominagb', 'nominagbv', 'nominagbf']
arr_keys = {
    'cuotas_csl': {
        'description': 'Descuento programado: Cooperativa San Luis',
        'nominas': {
            'nominagb': 81,
            'nominagbv': 82,
            'nominagbf': 100,
        },
    },
    'microcreditos': {
        'description': 'Descuento programado: Microcredito',
        'nominas': {
            'nominagb': 107,
            'nominagbv': 107,
            'nominagbf': 107,
        },
    },
    'cuotas_figua': {
        'description': 'Descuento programado: Cuota Prestamo Figua',
        'nominas': {
            'nominagb': 106,
            'nominagbv': 106,
            'nominagbf': 106,
        },
    },
    'cuotas_ahorro': {
        'description': 'Descuento programado: Ahorro Empresa',
        'nominas': {
            'nominagb': 112,
            'nominagbv': 112,
            'nominagbf': 112,
        },
    },
    'cuota_seguro': {
        'description': 'Descuento programado: Cuota de Seguro',
        'nominas': {
            'nominagb': 54,
            'nominagbv': 54,
            'nominagbf': 54,
        }
    },
    'cuota_jornadas': {
        'description': 'Descuento programado: Cuota de Vacunas, Jornadas y Bazares',
        'nominas': {
            'nominagb': 74,
            'nominagbv': 73,
            'nominagbf': 80,
        }
    },
    'farmacia': {
        'description': 'Descuento programado: Farmacia',
        'nominas': {
            'nominagb': 82,
            'nominagbv': 81,
            'nominagbf': 99,
        },
    },
    'cuotas_upa': {
        'description': 'Descuento programado: Cuota UPA',
        'nominas': {
            'nominagb': 78,
            'nominagbv': 78,
            'nominagbf': 92,
        },
    },
    'cuotas_otros': {
        'description': 'Descuento programado: Otros Descuentos',
        'nominas': {
            'nominagb': 61,
            'nominagbv': 61,
            'nominagbf': 61,
        },
    },
}


@login_required(login_url="/login/")
def index(request):
    data = { 'columns': columns }
    return render(request, 'carga_archivos_pagos/carga_archivos_pagos.html', data)


@login_required(login_url="/login/")
def validate_employee(request, value, row, int_fila_error):
    arr_result = []
    bool_error = False
    str_response = ''
    if str(value) == '' or str(value) == 'None' or not value:
        bool_error = True
        str_response += " la fila '%s' no tiene código de empleado, " % int_fila_error
    else:
        str_name_employee = row[int_key_name_column].value
        if str(str_name_employee) == '' or str(str_name_employee) == 'None' or not str_name_employee:
            bool_error = True
            str_response = """En la fila %s se encontro un empleado sin nombre valido""" % int_fila_error
        bool_employee_done = False
        if not bool_error:
            str_name_employee = str_name_employee.lower()
            str_name_employee = unidecode.unidecode(str_name_employee)
            for str_n in arr_nominas:
                if not bool_employee_done:
                    str_query_employee = """SELECT Nombres, Apellidos, No_Empresa, No_Puesto, DUI
                                                FROM %s..Empleados
                                            WHERE No_Empleado = '%s' AND fecha_baja IS NULL""" % (str_n, value)
                    arr_prev_employee = get_query(str_query_employee, True)

                    if len(arr_prev_employee) > 0:
                        arr_prev_name = arr_prev_employee[0]['Nombres'].split(' ')
                        arr_prev_lastname = arr_prev_employee[0]['Apellidos'].split(' ')
                        int_coincidence = 0
                        for str_name in arr_prev_name:
                            str_name = unidecode.unidecode(str_name)
                            int_result = str_name_employee.find(str_name.lower())
                            if int_result != -1:
                                int_coincidence += 1

                        for str_lastname in arr_prev_lastname:
                            str_lastname = unidecode.unidecode(str_lastname)
                            int_result = str_name_employee.find(str_lastname.lower())
                            if int_result != -1:
                                int_coincidence += 1

                        if int_coincidence < 2:
                            bool_error = True
                            str_response += """\n El empleado aparece en la nomina %s con un código que 
                                                no coincide con su nombre, fila No. %s""" % (str_n, int_fila_error)
                        else:
                            str_response = ""
                            bool_employee_done = True
                            bool_error = False
                            arr_prev_employee[0]['nomina'] = format(str_n)
                            arr_result = arr_prev_employee[0]
                    else:
                        bool_error = True

    if bool_error:
        str_response = """El codigo de empleado de la linea %s no aparece en ninguna nomina.""" % int_fila_error

    return { 'error': bool_error, 'str_response': str_response, 'result': arr_result }


@login_required(login_url="/login/")
def process_file(request):
    file_excel = request.FILES.get('excel')
    str_option = request.POST.get('options', '')
    wb = openpyxl.load_workbook(file_excel)
    sheets = wb.sheetnames
    worksheet = wb[sheets[0]]
    arr_columns_tmp = get_columns_by_filter(request)
    arr_result_empleado = []
    bool_error = False
    int_key = 0
    str_saldo = 0
    int_fila_error = 1
    int_columns_expected = len(arr_columns_tmp)
    str_columns = ""
    str_query = ""
    str_prev_error = ""

    for row in worksheet.iter_rows():
        int_cell = 0
        str_values = ""

        for cell in row:
            if int(int_cell) < int(int_columns_expected):
                list_detail_column = arr_columns_tmp[int_cell]
                column = list_detail_column['str_position']
                value = str(cell.value).lower()
                if not bool_error:
                    str_comma = ","
                    if int_key <= 0:
                        if str(value) != 'None':
                            if value != column:
                                bool_error = True
                                str_prev_error = "La columna '%s' no debe de llamarse asi" % value

                            if str_columns == "":
                                str_comma = ""
                            if value == 'codigo_empleado':
                                value = 'cod_empleado'

                            if value != 'nombres':
                                str_columns += "%s %s" % (str_comma, value)
                    else:
                        if column == 'codigo_empleado':
                            arr_validate = validate_employee(request, value, row, int_fila_error)
                            bool_error = arr_validate['error']
                            str_prev_error += arr_validate['str_response']
                            arr_result_empleado = arr_validate['result']
                            if int_key <= 1:
                                str_columns += ", nombres, apellidos, dpi, empresa, no_credito, nomina"
                        if column == 'cuota' or column == 'descuento':
                            value = value.replace('Q', '').replace(' ', '').replace(',', '')
                            str_saldo = value

                        if str(value) != 'None':
                            if str_values == "":
                                str_comma = ""

                            if column != 'nombres':
                                str_values += "%s '%s'" % (str_comma, value)
            int_cell += 1
        if len(arr_result_empleado) > 0:
            str_values += ", '%s', '%s', '%s', '%s', '%s', '%s'" % (arr_result_empleado['Nombres'],
                            arr_result_empleado['Apellidos'], arr_result_empleado['DUI'],
                            arr_result_empleado['No_Empresa'], '001', arr_result_empleado['nomina'])

        if not bool_error or ( str(str_values) != '' and str_values ):
            if str(str_values) != '':
                str_query += """INSERT INTO nova..rrhh_pagos_programados_empleados (%s, movido_nomina, saldo_pendiente,
                                tipo_movimiento, created_at, updated_at) VALUES (%s, %s, '%s', '%s', GETDATE(),
                                GETDATE()); \n""" % (str_columns, str_values, '0', str_saldo, str_option)
        int_key += 1
        int_fila_error += 1

    str_error = "Verifica tu información por favor, %s." % str_prev_error
    if not bool_error:
        bool_insert = insert_query(str_query)
        if bool_insert:
            set_notification(request, True, "Los descuentos se grabaron esperando a que los muevas a la nómina",
                             "add_alert", "success")
        else:
            set_notification(request, False, 'Archivo correcto, fallo al insertar los detalles, contacta con IT',
                             "warning", "danger")
    else:
        set_notification(request, False, str_error, "warning", "danger")

    return redirect("rrhh-carga_archivos")


@login_required(login_url="/login/")
def get_columns_by_filter(request):
    arr_return = []
    str_option = request.POST.get('options', '')
    for detail in columns:
        bool_exist = str_option in detail['apply']
        if bool_exist:
            arr_return.append(detail)
    return arr_return


@login_required(login_url="/login/")
def get_current_period(request):
    int_year = dt.date.today().year
    int_month = dt.date.today().month
    int_day = dt.date.today().day
    str_today = '%s%s%s' % (int_year, str(int_month).zfill(2), str(int_day).zfill(2))
    str_query = """SELECT TOP 1 Tipo_Periodo, No_Periodo, Fecha_Final, Fecha_Inicial FROM NominaGB..Periodos
                    WHERE Fecha_Inicial <= '%s' ORDER BY Fecha_Inicial DESC""" % str_today
    return get_query(str_query, True)


@login_required(login_url="/login/")
def process_cuotas_nomina(request, obj_period, str_nomina, int_clave, str_description, int_employee, int_amount):
    user_id = request.user.id
    type_period = 'Q'
    reference = '202012001'
    date_init = ''
    date_end = ''

    if len(obj_period) > 10:
        date_init = obj_period[0]['Fecha_Inicial'].replace('-', '')
        date_end = obj_period[0]['Fecha_Final'].replace('-', '')
    else:
        today_date = dt.datetime.now()
        day = today_date.day
        if day <= 15:
            date_init = today_date.replace(day=1)
            date_end = today_date.replace(day=15)
        else:
            type_period = 'M'
            date_init = today_date.replace(day=15)
            date_end = today_date.replace(day=calendar.monthrange(today_date.year, today_date.month)[1])
        date_init = date_init.strftime("%Y%m%d")
        date_end = date_end.strftime("%Y%m%d")

    str_query = """INSERT INTO %s..MovimientosProgramados (No_Empleado, No_Clave, Fecha_Inicio, Fecha_Final,
                    Total, Cuota, No_Pagos, Saldo, Observaciones, ID_Usuario, Tipo_Periodo) VALUES ('%s', '%s',
                    '%s', '%s', %s, %s, 0, %s, '%s', %s, '%s')""" % (str_nomina, int_employee,
                    int_clave, date_init, date_end, int_amount, int_amount, int_amount, str_description, user_id,
                    type_period)
    bool_response = execute_query(str_query)
    return bool_response


@login_required(login_url="/login/")
def get_movements_pending(request):
    set_query = """SELECT * FROM nova..rrhh_pagos_programados_empleados WHERE movido_nomina = 0"""
    pagos = get_query(set_query)
    arr_return = []
    for pago in pagos:
        if pago['deudor'] is None:
            str_deudor = '%s %s' % (pago['nombres'], pago['apellidos'])
        else:
            str_deudor = format(pago['deudor'])
        arr_return.append({
            'id': format(pago['id']),
            'cod_empleado': format(pago['cod_empleado']),
            'deudor': format(str_deudor),
            'cuota': format(pago['cuota']),
            'saldo_pendiente': format(pago['saldo_pendiente']),
            'created_at': format(pago['created_at']),
            'tipo_movimiento': format(pago['tipo_movimiento']),
            'nomina': format(pago['nomina']),
        })

    data = {
        'movements': arr_return,
    }
    return render(request, 'carga_archivos_pagos/carga_archivos_pagos_movimientos.html', data)


@login_required(login_url="/login/")
def guardar_nomina(request):
    arr_deudores = request.POST.getlist('deudor[]', None)
    arr_int_descontar = request.POST.getlist('descontar[]', None)
    arr_int_pendiente = request.POST.getlist('pendiente[]', None)
    arr_tipo_movimiento = request.POST.getlist('tipo_movimiento[]', None)
    int_key_deudor = 0
    bool_status = True
    str_return = 'Pagos movidos correctamente a nomina'
    obj_period = get_current_period(request)

    for deudor in arr_deudores:
        if bool_status:
            int_descontar = arr_int_descontar[int_key_deudor]
            int_pendiente = arr_int_pendiente[int_key_deudor]
            str_tipo_movimiento = arr_tipo_movimiento[int_key_deudor]
            bool_movido = '1'
            if float(int_pendiente) > 0:
                bool_movido = '0'

            str_query_existente = """SELECT id, tipo_movimiento, nomina FROM nova..rrhh_pagos_programados_empleados
                                        WHERE cod_empleado = '%s' AND movido_nomina = '0'
                                            AND tipo_movimiento = '%s'""" % (deudor, str_tipo_movimiento)
            arr_existente = get_query(str_query_existente)

            if len(arr_existente) > 0:
                str_option = arr_existente[0]['tipo_movimiento']
                str_nomina = arr_existente[0]['nomina']
                str_description = ''
                int_key = 0

                if str_option in arr_keys:
                    str_description = arr_keys[str_option]['description']
                    if str_nomina in arr_keys[str_option]['nominas']:
                        int_key = arr_keys[str_option]['nominas'][str_nomina]

                bool_nomina = process_cuotas_nomina(request, obj_period, arr_existente[0]['nomina'],
                                                    int_key, str_description, deudor, int_descontar)
                if not bool_nomina:
                    bool_status = False
                    str_return = 'Ocurrió un error al mover a nomina el empleado: "%s"' % deudor
                else:
                    str_query_update = """UPDATE nova..rrhh_pagos_programados_empleados
                                                    SET saldo_pendiente = '%s', movido_nomina = '%s'
                                                    WHERE id = '%s'""" % (
                    int_pendiente, bool_movido, arr_existente[0]['id'])
                    bool_insert = execute_query(str_query_update)
                    if not bool_insert:
                        bool_status = False
                        str_return = 'Ocurrió un error al actualizar los pagos programados del empleado: "%s"' % deudor
            else:
                bool_status = False
                str_return = 'Ocurrió un error, por que se inserto un usuario que ya no existe: "%s"' % deudor
            int_key_deudor += 1

    data = {
        'status': bool_status,
        'message': str_return,
    }
    return JsonResponse(data, safe=True)


@login_required(login_url="/login/")
def borrar_registros(request):
    arr_movements = request.POST.getlist('movements[]', None)
    bool_no_error = True
    str_message = "Borrado correctamente."

    for movement in arr_movements:
        if bool_no_error:
            str_query = """DELETE FROM nova..rrhh_pagos_programados_empleados WHERE id = '%s'""" % movement
            # response = False
            response = execute_query(str_query)
            if not response:
                bool_no_error = False
                str_message = "Ocurrió un error al borrar los registros, contacta con IT."

    data = {
        'status': bool_no_error,
        'message': str_message,
    }
    return JsonResponse(data, safe=True)


@login_required(login_url="/login/")
def borrar_registro(request):
    movement = request.POST.get('movement', '')
    bool_no_error = True
    str_message = "Borrado correctamente."
    str_query = """DELETE FROM nova..rrhh_pagos_programados_empleados WHERE id = '%s'""" % movement
    response = execute_query(str_query)
    if not response:
        bool_no_error = False
        str_message = "Ocurrió un error al borrar el registro."

    data = {
        'status': bool_no_error,
        'message': str_message,
    }
    return JsonResponse(data, safe=True)
